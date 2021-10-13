import openpyxl as op
import yaml
import sys
import ipcalc,math

##########################
# Global settings
##########################
debug=0
myyaml={}
mysecretyaml={}

##########################
# Open workbook
##########################
if (len(sys.argv) == 1):
  print("Please provide a spreadsheet file.")
  exit()
xlsxfile=sys.argv[1]
wb = op.load_workbook(filename = xlsxfile,read_only=True, data_only=True)

##########################
# DEBUG PRINT FUNCTION
##########################
def debugprint(message):
    if (debug == 1):
        print(message)


##########################
# Sheet Allocation for hosts
##########################
generalSheet = wb["General"]
generalSheetStart='A1'
generalSheetStop='B'

ws=wb['General']
for cnt, row in enumerate(generalSheet[generalSheetStart:generalSheetStop + str(generalSheet.max_row)]):
    if row[0].value is not None and row[0].value != 0:
        if row[0].value == "Tenant Name":
            myyaml['name'] = str(row[1].value)
            debugprint(myyaml['name'])
        elif row[0].value == "Environment":
            mysecretyaml['environment'] = str(row[1].value)
            debugprint(mysecretyaml['environment'])
#        elif row[0].value == "Site Name":
#            myyaml['site'] = str(row[1].value)
#            debugprint(myyaml['site'])
#        elif row[0].value == "Application Environment":
#            myyaml['applicationenvironment'] = str(row[1].value)
#            debugprint(myyaml['applicationenvironment'])
#        elif row[0].value == "Class of service":
#            myyaml['classofservice'] = str(row[1].value)
#            debugprint(myyaml['classofservice'])
#        elif row[0].value == "Boot Volume Size":
#            myyaml['bootvolumesize'] = str(row[1].value)
#            debugprint(myyaml['bootvolumesize'])
#        elif row[0].value == "Windows Image Name":
#            myyaml['windowsimagename'] = str(row[1].value)
#            debugprint(myyaml['windowsimagename'])
#        elif row[0].value == "RHEL Image Name":
#            myyaml['rhelimagename'] = str(row[1].value)
#            debugprint(myyaml['rhelimagename'])
#        elif row[0].value == "Ignored Server Types":
#            myyaml['ignoredservertypes'] = str(row[1].value)
#            debugprint(myyaml['ignoredservertypes'])
#        elif row[0].value == "Flavour Suffix":
#            myyaml['flavorsuffix'] = str(row[1].value)
#            debugprint(myyaml['flavorsuffix'])
#        elif row[0].value == "Default Permissions":
#            myyaml['defaultpermissions'] = str(row[1].value)
#            debugprint(myyaml['defaultpermissions'])
#        elif row[0].value == "shmmni (bytes)":
#            myyaml['shmni'] = str(row[1].value)
#            debugprint(myyaml['shmni'])
#        elif row[0].value == "Tenant State":
#            myyaml['tenantstate'] = str(row[1].value)
#            debugprint(myyaml['tenantstate'])
        elif row[0].value == "Description":
            myyaml['description'] = str(row[1].value)
            debugprint(myyaml['description'])
        elif row[0].value == "Instances":
            myyaml['instances'] = str(row[1].value)
            debugprint(myyaml['instances'])
        elif row[0].value == "Storagesize":
            myyaml['storagesize'] = str(row[1].value)
            debugprint(myyaml['storagesize'])
        elif row[0].value == "Memory":
            myyaml['memory'] = str(row[1].value)
            debugprint(myyaml['memory'])
        elif row[0].value == "Volumes":
            myyaml['volumes'] = str(row[1].value)
            debugprint(myyaml['volumes'])
        elif row[0].value == "CPUS":
            myyaml['cpus'] = str(row[1].value)
            debugprint(myyaml['cpus'])
#print(myyaml)
print(xlsxfile)
print(yaml.dump({'project': myyaml}, explicit_start=True,default_flow_style=False,sort_keys=False))
with open('tenant-project-config_'+mysecretyaml['environment'].lower()+'.yml', 'w') as yaml_file:
    yaml.dump({'project': myyaml}, yaml_file, explicit_start=True,default_flow_style=False,sort_keys=False)


##########################
# Sheet Allocation for network
##########################
generalSheet = wb["Security Zones"]
generalSheetStart='A8'
generalSheetStop='K'

def gateway(gwcidr):
    netdata=ipcalc.Network(gwcidr)
    netdatalist=list(netdata)
    netdatalistlen=len(netdatalist)
    return str(netdatalist[0])
    
def iprange(rangecidr):
    netdata=ipcalc.Network(rangecidr)
    netdatalist=list(netdata)
    netdatalistlen=len(netdatalist)
    amount=0
    if (netdatalistlen in [6,14,30]):
        amount=4
    if (netdatalistlen > 30):
        amount=math.ceil(netdatalistlen/10)
    return [str(netdatalist[netdatalistlen-amount]), str(netdatalist[netdatalistlen-1])]
    
ws=wb['General']
data={}
envfound=0
envname=''
for cnt, row in enumerate(generalSheet[generalSheetStart:generalSheetStop + str(generalSheet.max_row)]):
    if envfound==0 and row[0].value is not None and row[0].value != 0:
        debugprint(data)
        envfound=1
        envname=row[0].value
        cidr=row[2].value+row[3].value
        subnet=ipcalc.Network(cidr).network()
        route_asn=row[6].value
        route_target=row[7].value
        gw=gateway(cidr)
        range=iprange(cidr)
        data={envname : [{'cidr': cidr, 'name': row[1].value, 'route_asn': row[6].value, 'route_target': row[7].value, 'subnet': str(subnet),'subnet_prefix': int(row[3].value.replace('/','')),'gateway':gw,'allocation_start':range[0],'allocation_end':range[1]}]}
        current_environment=str(row[0].value)
        current_network=str(row[1].value)
    if envfound != 0 and row[0].value is not None and row[0].value != 0 and row[2].value is not None and row[2].value != 0 and current_environment == str(row[0].value) and current_network != str(row[1].value):
        debugprint(data)
        #envname=row[0].value
        cidr=row[2].value+row[3].value
        subnet=ipcalc.Network(cidr).network()
        route_asn=row[6].value
        route_target=row[7].value
        gw=gateway(cidr)
        range=iprange(cidr)
        #data.update({envname: []})
        data[envname].append({'cidr': cidr, 'name': row[1].value, 'route_asn': row[6].value, 'route_target': row[7].value, 'subnet': str(subnet),'subnet_prefix': int(row[3].value.replace('/','')),'gateway':gw,'allocation_start':range[0],'allocation_end':range[1]})
        current_environment=str(row[0].value)
        current_network=str(row[1].value)
    if envfound != 0 and row[0].value is not None and row[2].value is not None and row[2].value != 0 and current_environment != str(row[0].value) and current_network != str(row[1].value):
        envname=row[0].value
        cidr=row[2].value+row[3].value
        subnet=ipcalc.Network(cidr).network()
        route_asn=row[6].value
        route_target=row[7].value
        gw=gateway(cidr)
        range=iprange(cidr)
        data.update({envname: []})
        data[envname].append({'cidr': cidr , 'name': row[1].value, 'route_asn': row[6].value, 'route_target': row[7].value, 'subnet': str(subnet),'subnet_prefix': int(row[3].value.replace('/','')),'gateway':gw,'allocation_start':range[0],'allocation_end':range[1]})
        current_environment=str(row[0].value)
        current_network=str(row[1].value)
debugprint(data)
#print(yaml.dump({'tenant_networks': data}, explicit_start=True,default_flow_style=False,sort_keys=False))

for singleenv in data:
  with open('tenant-network-config_'+singleenv.lower()+'.yml', 'w') as yaml_file:
      yaml.dump({'tenant_networks': data[singleenv]}, yaml_file, explicit_start=True,default_flow_style=False,sort_keys=False)
