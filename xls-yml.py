#!/usr/bin/env python
import yaml
import json
import os
import csv
import codecs
import sys
import pathlib
import argparse
import logging
import textwrap
import requests
import datetime
import getpass
import copy
import distutils
import openpyxl
from openpyxl import load_workbook


#from funks.dev_create_csr_json import * # Imports all functions within file natively, therefore cannot have identically named functions
import funks # Imports all functions as defined in package funks/__init__.py
             # Functions can be called as: funks.pretty_json() or funks.create_csr_json.pretty_json()

logger = logging.getLogger(__name__)
sh = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
sh.setFormatter(formatter)
logger.addHandler(sh)


def validate_csv_file(csvFile):

   #csvSeek = codecs.open(LBSheet, "r", 'utf-8')
    #csvOpen = csv.reader(LBSheet)   # by passing this we're no longer streaming the file, but we need to read lines
    #keys = next(csvOpen) # next(csvSeek) to revert to streaming the file
    #sheet_obj = LBSheet.active
    #csvOpen = 
    LBSheet = csvFile["LB"]
    max_col = LBSheet.max_column
    max_row = LBSheet.max_row
    keys = []
    for i in range(1, max_col + 1):
      cell_obj = LBSheet.cell(row = 3, column = i)
      key= cell_obj.value
      keys.append(key)
    fullrow_value= []
    for x in range (4, max_row + 1):
      onerow_value = []
      if LBSheet.cell(row = x, column = 1).value:
        for i in range(1, max_col + 1):
          cell_obj = LBSheet.cell(row = x, column = i)
          cell_value= cell_obj.value
          onerow_value.append(cell_value)
        fullrow_value.append(onerow_value)
    dict_per_app = [] # list of dicts for each 'app' containing pool:vsvip:virtualservice
    js_current = {} # Current dict
    js_next = {} # Next dict
    csrs_required = {}
    tenant = ""
    csr_names = []
    persistence_profiles = {}
    controller = ""
    generate_please = False
    for row in fullrow_value:
        slb_data = dict(zip(keys, row))
        if slb_data['dc'] != 'a' and slb_data['dc'] != 'b' and slb_data['dc'] != 'dev_a' and slb_data['dc'] != 'dev_b' and slb_data['dc'] != 'preprod_a' and slb_data['dc'] != 'preprod_b': 
            print ("Wrong datacentre. Correct the CSV file.")
            exit()
        if slb_data['dc'] == 'a':
            controller = "10.126.163.8"
        elif slb_data['dc'] == 'b':
            controller = "10.124.163.8"
        elif slb_data['dc'] == 'dev_a':
            controller = "10.244.174.8"    
        elif slb_data['dc'] == 'dev_b':
            controller = "10.246.174.8"
        elif slb_data['dc'] == 'preprod_a':
            controller = "10.126.163.16"
        elif slb_data['dc'] == 'preprod_b':
            controller = "10.124.163.16"
        else:
            print("Invalid controller. Exiting")
            sys.exit()
        tenant =  slb_data['tenant']
        if 'System-Secure-HTTP' in str(slb_data['application_profile_ref']) or 'SSL' in slb_data['application_profile_ref']:
            if str(slb_data['ssl_key_and_certificate_refs']) == "":# or str(slb_data['ssl_key_and_certificate_refs']) == "None":
                csr_data = {}
                csr_data['dns_name'] = slb_data['fqdn']
                csr_data['friendly_name'] = slb_data['gslb_cname']
                csr_data['short_name'] = slb_data['vs_name']
                csr_data['tenant'] = slb_data['tenant']
                csr_data['controller'] = slb_data['dc']
                if csr_data not in csr_names:
                    csr_names.append(csr_data)
        if slb_data['application_persistence_profile_ref'] != "None" and slb_data['application_persistence_profile_ref'] != "":
            persistence_profiles[slb_data['application_persistence_profile_ref']] = tenant

    # username="aleksy"
    # print ("Enter password for user " + username+":")
    # password=getpass.getpass()      
    # api = requests.session()
    # api.verify = False
    # api.headers['x-avi-tenant'] = tenant
    # r = api.post('https://' + controller + '/login', json={'username': username, 'password': password})
    # r.raise_for_status()
    # api.headers['referer'] = 'https://' + controller
    # api.headers['x-csrftoken'] = api.cookies['csrftoken']
    # r = api.get( 'https://'+controller+'/api/applicationpersistenceprofile')
    # for profile in persistence_profiles:
    #     for result in r.json()['results']:
    #         print (result)
    #         if profile in result['name']:
    #             print(profile+' not required')
    
    if csr_names:
        p = pathlib.Path(__file__).resolve().parent.parent
        parent = p.resolve()
        csr_base_path = parent.joinpath("applications/"+tenant+"/csrs")
        if (csr_base_path.exists() == False):
            csr_base_path.mkdir(parents=True, exist_ok=True) # mkdir
        tenant_dir = tenant
        if (csr_base_path.exists() == False):
            csr_base_path.mkdir(parents=True, exist_ok=True) # mkdir
        else:
            logger.info("Directory: %s exists", dir)
    
        #Creating a json file that can be used to generate CSRs
        pathlib.Path

        json_file = pathlib.Path(str(csr_base_path)+"/names.json")
        if json_file.exists ():
            print ('names.json exists')
            with open(str(csr_base_path)+"/names.json") as f:
                names_in_file = f.read()
            print (names_in_file)
            print (json.dumps(csr_names))
            if names_in_file == json.dumps(csr_names):
                print ("No new csrs required "+str(csr_base_path)+" directory")
                generate_please = False
            else:
                json_file = open(str(csr_base_path)+"/names.json", "w")
                json_file.write(json.dumps(csr_names))
                json_file.close()
                print ("New json file for csrs has been created in "+str(csr_base_path)+" directory")
                generate_please = True
        else:
            json_file = open(str(csr_base_path)+"/names.json", "w")
            json_file.write(json.dumps(csr_names))
            json_file.close()
            print ("New json file for csrs has been created in "+str(csr_base_path)+" directory")
            generate_please = True

    if generate_please:
            generate = input("Do you want to generate CSRs now? (y/n)")
            if generate.lower() == 'y':
                try:
                    username=os.environ['OS_USERNAME']
                except:
                    username=input("Username not found, enter username:")
                try:
                    password = os.environ['OS_PASSWORD']
                except:
                    print ("Enter password for user" + username+":")
                    password=getpass.getpass()

                now = datetime.datetime.now()
                today = datetime.date.today()
                base_domain = 'uk.pri.o2.com'
                csr_template = {
                "type": "SSL_CERTIFICATE_TYPE_VIRTUALSERVICE",
                "certificate": {
                    "days_until_expire": 30,
                    "self_signed": "false",
                    "subject": {
                    "locality": "Slough",
                    "country": "GB",
                    "state": "England",
                    "organization": "TELEFONICA UK LIMITED",
                    "organization_unit": "Operations"
                    }
                }
                }


                ##x-avi-version 17.2.14
                with open(str(csr_base_path)+"/names.json") as fh:
                    names = json.load(fh)
                for name in names:          
                    ctrl = name['controller']
                    if ctrl == 'a':
                        controller = "10.126.163.8"
                    elif ctrl == 'b':
                        controller = "10.124.163.8"
                    elif ctrl == 'dev':
                        controller = "10.244.174.8"    
                    elif ctrl == 'dev_b':
                        controller = "10.246.174.8"
                    elif ctrl == 'preprod_a':
                        controller = "10.126.163.16"
                    elif ctrl == 'preprod_b':
                        controller = "10.124.163.16"
                    else:
                        print("Invalid controller. Exiting")
                        sys.exit()
                    api = requests.session()
                    api.verify = False
                    r = api.post('https://' + controller + '/login', json={'username': username, 'password': password})
                    r.raise_for_status()
                    api.headers['referer'] = 'https://' + controller
                    api.headers['x-csrftoken'] = api.cookies['csrftoken']
                    data = copy.deepcopy(csr_template)
                    data['name'] = '%s-%s' % (name['dns_name'], today)
                    data['certificate']['subject']['common_name'] = name['dns_name']
                    san = []
                    san.append(name['dns_name'])
                    san_fn = name.get('friendly_name', "None") # friendly_name here is gslb_cname in csv
                    if san_fn != "None":
                        if isinstance(san_fn, list):
                            for i in range(len(san_fn)):
                                san.append(san_fn[i])
                        else:
                            san.append(name['friendly_name'])
                    data['certificate']['subject_alt_names'] = san
                    api.headers['x-avi-tenant'] = name['tenant']

                    r = api.post( 'https://'+controller+'/api/sslkeyandcertificate', json=data)
                    try:
                        r.raise_for_status()
                    except requests.exceptions.HTTPError as e:
                        if str(e).split()[0] == '500':
                            if data['certificate']['subject_alt_names'][-1] == '':
                                data['certificate']['subject_alt_names'].pop()
                                r = api.post( 'https://'+controller+'/api/sslkeyandcertificate', json=data)
                                r.raise_for_status()
                    with open(str(csr_base_path)+'/%s.pem' % data['name'], 'w') as fh:
                        try:
                            fh.write(r.json()['certificate']['certificate_signing_request'])
                        except KeyError as e:
                            print ("[WARN]: PEM data not in creation response: %s" % r.text)
                            r = api.get(r.json()['url'])
                            r.raise_for_status()
                            fh.write(r.json()['certificate']['certificate_signing_request'])

# Create Pools List of Dicts from the CSV
def create_pools_lod_from_csv(csvFile):
    #csvSeek = codecs.open(csvFile, "r", 'utf-8')
    #csvOpen = csv.reader(csvSeek)   # by passing this we're no longer streaming the file, but we need to read lines
    #keys = next(csvOpen) # next(csvSeek) to revert to streaming the file
    LBSheet = csvFile["LB"]
    max_col = LBSheet.max_column
    max_row = LBSheet.max_row
    keys = []
    for i in range(1, max_col + 1):
      cell_obj = LBSheet.cell(row = 3, column = i)
      key= cell_obj.value
      keys.append(key)
    fullrow_value= []
    for x in range (4, max_row + 1):
      onerow_value = []
      if LBSheet.cell(row = x, column = 1).value:
        for i in range(1, max_col + 1):
          cell_obj = LBSheet.cell(row = x, column = i)
          cell_value= cell_obj.value
          onerow_value.append(cell_value)
        fullrow_value.append(onerow_value)
    dict_per_app = [] # list of dicts for each 'app' containing pool:vsvip:virtualservice
    current_pool = "None"
    pool = "None"
    js_current = {} # Current dict
    js_next = {} # Next dict
    for row in fullrow_value:
        d = dict(zip(keys, row))
        create=d.get('create', 'None')
        pool = d['vs_name']+'-pool'
        if str(create).lower() == 'true' or  create == 'None' or str(create).lower() == 'yes':              
            if current_pool != d['vs_name']+'-pool':
                js_pool = {}                            # Create a new object            
                js_pool['name'] = d['vs_name']+'-pool'        # Add data
                js_pool['enabled'] = str(d['enabled']).lower()
                js_pool['tenant'] = d['tenant']
                js_pool['cloud_ref'] = '/api/cloud?name=' + d['cloud_ref']
                js_pool['state'] = d['state']
                js_pool['lb_algorithm'] = d['lb_algorithm']
                js_pool['lb_algorithm_hash'] = d['lb_algorithm_hash']
                if d['application_persistence_profile_ref']:
                    js_pool['application_persistence_profile_ref'] = '/api/applicationpersistenceprofile?name=' + d['application_persistence_profile_ref']
                health = []    
                if d['health_monitor_refs'] == 'Customised-HTTP':
                    health.append('/api/healthmonitor?name='+ d['vs_name'] + '_hm')
                else:
                    health_monitor_refs = d['health_monitor_refs'].split(",") # 
                    for line in health_monitor_refs:
                       health.append('/api/healthmonitor?name='+ line)
                js_pool['health_monitor_refs'] = health

                # This is a better way of evaluating if a key is available in the dict, as opposed to throwing a KeyError and catching it
                # Using this method, can avoid backporting the field into every CSV/XLSX
                js_pool['sni_enabled'] = d.get('pool_sni_enabled', "false") # Default returns NoneType, so setting a string to "None" to strip later
                if js_pool['sni_enabled'] == '':
                    js_pool['sni_enabled'] = 'false'
                elif str(js_pool['sni_enabled']).lower() == 'true' or js_pool['sni_enabled'] == 'yes':
                    js_pool['ssl_profile_ref'] = '/api/sslprofile?name=' + d.get('pool_ssl_profile_ref', "System-Standard")
                    js_pool['ssl_to_backend'] = 'true'
                # Add Servers to Pool Object
                js_pool['servers'] = []
                all_ports =[]
                if '/' in str(d['port']):   
                    all_ports=d['port'].split('/')
                elif '-' in str(d['port']):
                    first_port,last_port=d['port'].split('-')
                    for port in range(int(first_port), int(last_port)+1):
                        all_ports.append(str(port))
                else:
                    all_ports.append(d['port'])

                for current_port in all_ports:
                    js_servers = {}
                    js_servers['enabled'] = str(d['svr_enabled']).lower()
                    js_servers['hostname'] = d['hostname']
                    js_servers['ip'] = {}
                    js_servers['ip']['addr'] = d['addr']
                    if js_servers['ip']['addr'] == "":
                        js_servers['resolve_server_by_dns'] = 'true'
                        js_servers['ip']['addr'] = ""
                        js_servers['ip']['type'] = 'V4'
                    elif ':' in js_servers['ip']['addr']:
                        js_servers['ip']['type'] = 'V6'
                    elif '.' in js_servers['ip']['addr']:
                        js_servers['ip']['type'] = 'V4'
                    js_servers['port'] = current_port
                    js_servers['verify_network'] = str(d['verify_network']).lower()
                    js_servers['description'] = d['description']
                    js_pool['servers'].append(js_servers)
                # Logic to write-out to array if we've moved onto next line in the csv
                #if current_pool != d['vs_name']+'-pool': # Evalutes to true as current_pool is last but one
                #js_last = js_pool
                if bool(js_next):
                    dict_per_app.append(js_next)        # Write out last object to arai when js_pool is new
                    js_next = {}

                if bool(js_current):
                    if js_current['name'] != d['vs_name']+'-pool':
                        dict_per_app.append(js_current)

                js_current = js_pool
                current_pool = d['vs_name']+'-pool'
            elif current_pool == d['vs_name']+'-pool':
                all_ports =[]
                if '/' in str(d['port']):   
                    all_ports=str(d['port']).split('/')
                elif '-' in str(d['port']):
                    first_port,last_port=str(d['port']).split('-')
                    for port in range(int(first_port), int(last_port)+1):
                        all_ports.append(str(port))
                else:
                    all_ports.append(str(d['port'])) 
                for current_port in all_ports:
                    js_servers = {}
                    js_servers['enabled'] = str(d['svr_enabled']).lower()
                    js_servers['hostname'] = d['hostname']
                    js_servers['ip'] = {}
                    js_servers['ip']['addr'] = d['addr']
                    if js_servers['ip']['addr'] == "":
                        js_servers['resolve_server_by_dns'] = 'true'
                        js_servers['ip']['addr'] = ""
                        js_servers['ip']['type'] = 'V4'
                    elif ':' in js_servers['ip']['addr']:
                        js_servers['ip']['type'] = 'V6'
                    elif '.' in js_servers['ip']['addr']:
                        js_servers['ip']['type'] = 'V4'
                    js_servers['port'] = current_port
                    js_servers['verify_network'] = str(d['verify_network']).lower()
                    js_servers['description'] = d['description']
                    js_pool['servers'].append(js_servers)                
                pool = d['vs_name']+'-pool'
                js_next = js_pool
                js_current = {} # Zero the dict  
        else:
            continue
    dict_per_app.append(js_pool) # Final write of object in memory to arai - the last line of the csv
    return dict_per_app

# Create Virtual Services List of Dicts from the CSV
def create_vs_lod_from_csv(csvFile):
    #csvSeek = codecs.open(csvFile, "r", 'utf-8')
    #csvOpen = csv.reader(csvSeek)   # by passing this we're no longer streaming the file
    #keys = next(csvOpen) # next(csvSeek) to revert to streaming the file
    LBSheet = csvFile["LB"]
    max_col = LBSheet.max_column
    max_row = LBSheet.max_row
    keys = []
    for i in range(1, max_col + 1):
      cell_obj = LBSheet.cell(row = 3, column = i)
      key= cell_obj.value
      keys.append(key)
    fullrow_value= []
    for x in range (4, max_row + 1):
      onerow_value = []
      if LBSheet.cell(row = x, column = 1).value:      
        for i in range(1, max_col + 1):
          cell_obj = LBSheet.cell(row = x, column = i)
          cell_value= cell_obj.value
          onerow_value.append(cell_value)
        fullrow_value.append(onerow_value)
    dict_per_app = [] # list of dicts for each 'app' containing pool:vsvip:virtualservice
    current_row = None
    last_row = None
    js_current = {} # Current dict
    js_next = {} # Next dict
    for row in fullrow_value:
        d = dict(zip(keys, row))
        create=d.get('create', 'false')
        if str(create).lower() == 'true' or  create == 'None' or str(create).lower() == 'yes':
            if current_row != d['vs_name']:
                js_vs = {}                            # Create a new object            
                js_vs['name'] = d['vs_name']        # Add data
                js_vs['enabled'] = str(d['vs_enabled']).lower()
                js_vs['tenant'] = d['tenant']
                js_vs['cloud_ref'] = '/api/cloud?name=' + d['cloud_ref']
                js_vs['state'] = d['state']
#                js_vs['vsvip_ref'] = d['vsvip_ref']
                #js_vs['fqdn'] = d['fqdn']
                #js_vs['dns_info'] = []
                #js_dns_info = { "type": {}, "ttl": {}, "fqdn": {} }
                # Hardcode some values as no time to update csv
                #js_dns_info_type = "DNS_RECORD_A"
                #js_dns_info_ttl = "30"
                # Add to dict
                #js_dns_info['type'] = js_dns_info_type
                #js_dns_info['ttl'] = js_dns_info_ttl
                #js_dns_info['fqdn'] = d['fqdn']
                #js_vs['dns_info'].append(js_dns_info)
                js_vs['vip'] = []
                js_vsvip = {}
                js_vsvip['vip_id'] = d['vs_name']
                static_ip = d.get('ip_addr', "None") # Added static_ip for BSS_Migration1
                if static_ip != "None" and static_ip != "":
                    js_vsvip['ip_address'] = { "addr": {}, "type": {} }
                    js_vsvip['ip_address']['addr'] = d.get('ip_addr', "None")
                    js_vsvip['ip_address']['type'] = 'V4'
                    js_vsvip['auto_allocate_ip'] = 'false'
                else:
                    js_vsvip['auto_allocate_ip'] = 'true'              
                    vip_subnet_cidr = d.get('vip_subnet_cidr', 'None')
                    if vip_subnet_cidr != 'None':
                        vs_subnet,vs_mask = vip_subnet_cidr.split('/')
                        js_vsvip['subnet'] = {}
                        js_vsvip['subnet']['ip_addr'] ={}
                        js_vsvip['subnet']['ip_addr']['addr'] = vs_subnet
                        js_vsvip['subnet']['mask'] = vs_mask
                        js_vsvip['subnet']['ip_addr']['type'] = 'V4'
                    else:
                        js_vsvip['subnet_uuid'] = d.get('vip_subnet_uuid', 'None')
                js_vs['vip'].append(js_vsvip)
                js_vs['se_group_ref'] = '/api/serviceenginegroup?name=' + d['se_group_ref']

                js_vsds={}
                index=1
                if d.get('datascripts') != None and str(d.get('datascripts')).lower() != 'none':
                    js_vs['vs_datascripts'] = []
                    if '/' in d['datascripts']:   
                        for datascript in d['datascripts'].split('/'):
                            js_vsds['index'] = str(index)
                            js_vsds['vs_datascript_set_ref'] = '/api/vsdatascriptset/?name='+datascript
                            js_vs['vs_datascripts'].append(js_vsds)
                            index = ++index
                    elif ';' in d['datascripts']:
                        for datascript in d['datascripts'].split(';'):
                            js_vsds['index'] = str(index)
                            js_vsds['vs_datascript_set_ref'] = '/api/vsdatascriptset/?name='+datascript
                            js_vs['vs_datascripts'].append(js_vsds)
                            index = ++index               
                    else:
                            js_vsds['index'] = str(index)
                            js_vsds['vs_datascript_set_ref'] = '/api/vsdatascriptset/?name='+d['datascripts']
                            js_vs['vs_datascripts'].append(js_vsds)
                
                js_vs['services'] = []
                js_vs_services = {}
                # This needs updating to handle multiple ports. Each port requires tuple of (enable_ssl, port)
                js_vs_services['port'] = d['services_port'] #.split(",")
                js_vs['services'].append(js_vs_services)              
                js_vs['application_profile_ref'] = '/api/applicationprofile?name=' + d['application_profile_ref']
                if 'Secure-HTTP' in d['application_profile_ref'] or 'SSL' in d['application_profile_ref']:
                    js_vs_services['enable_ssl'] = 'true' #.split(",")
                    js_vs['ssl_key_and_certificate_refs'] = '/api/sslkeyandcertificate?name=' + str(d.get('ssl_key_and_certificate_refs', 'System-Default-Cert')) # Passing a single value instead of a list works and is tripped when 'None' is specified in csv
                    if d['ssl_key_and_certificate_refs'] == '':
                        js_vs['ssl_key_and_certificate_refs'] = '/api/sslkeyandcertificate?name=System-Default-Cert'
                    js_vs['ssl_profile_ref'] = '/api/sslprofile?name=' + d['ssl_profile_ref']
                    
                #ssl_key_and_certificate_refs = []
                #ssl_key_and_certificate_refs.append('/api/sslkeyandcertificate?name=' + d['ssl_key_and_certificate_refs'])
                #js_vs['ssl_key_and_certificate_refs'] = ssl_key_and_certificate_refs  # This can be multi-value? .split(",") - Possible fix in place for health_monitor and 'None' value
                js_vs['pool_ref'] = '/api/pool?name=' + d['vs_name']+'-pool' # Use this istead of repeating value unless you wanted to change pools later on?
                if d['http_policy_set_ref'] != "None" and d['http_policy_set_ref'] != "":  #  This goes against using my "None" key removal, but keyremoval won't work on the higher level key or http_policies which returns blank. Therefore this is a bit of a hack
                    js_vs['http_policies'] = []
                    #refs = d['http_policy_set_ref'].split(";")
                    js_href = {}
                    #js_href['http_policy_set_ref'] = [] 
                    #for r in refs:
                    js_href['http_policy_set_ref'] = ('/api/httppolicyset?name=' + d['http_policy_set_ref'])
                    js_vs['http_policies'].append(js_href)
                if current_row and last_row != d['vs_name']:
                    dict_per_app.append(js_current)        # Write out object to arai
                js_current = js_vs # Set newly created object to spare variable
                current_row = d['vs_name']
            elif current_row == d['vs_name']:
                last_row = d['vs_name'] 
        else:
            continue
    dict_per_app.append(js_current) # Dump last row
    return dict_per_app

def create_cert_load_from_csv(csvFile):
    #csvSeek = codecs.open(csvFile, "r", 'utf-8')
    #csvOpen = csv.reader(csvSeek)   # by passing this we're no longer streaming the file
    #keys = next(csvOpen) # next(csvSeek) to revert to streaming the file
    LBSheet = csvFile["LB"]
    max_col = LBSheet.max_column
    max_row = LBSheet.max_row
    keys = []
    for i in range(1, max_col + 1):
      cell_obj = LBSheet.cell(row = 3, column = i)
      key= cell_obj.value
      keys.append(key)
    fullrow_value= []
    for x in range (4, max_row + 1):
      onerow_value = []
      if LBSheet.cell(row = x, column = 1).value:      
        for i in range(1, max_col + 1):
          cell_obj = LBSheet.cell(row = x, column = i)
          cell_value= cell_obj.value
          onerow_value.append(cell_value)
        fullrow_value.append(onerow_value)
    dict_per_app = [] # list of dicts for each 'app' containing pool:vsvip:virtualservice
    current_row = None
    last_row = None
    js_current = {} # Current dict
    js_next = {} # Next dict
    for row in fullrow_value:
        d = dict(zip(keys, row))
        create=d.get('create', 'false')
        if str(create).lower() == 'true' or  create == 'None' or str(create).lower() == 'yes':
            if current_row != d['vs_name']:
                if 'Secure-HTTP' in d['application_profile_ref'] or 'SSL' in d['application_profile_ref']:
                  js_cert = {}                            # Create a new object            
                  js_cert['name'] = d['ssl_key_and_certificate_refs']        # Add data
                  js_cert['tenant'] = d['tenant']
                  js_cert['cloud_ref'] = '/api/cloud?name=' + d['cloud_ref']
                  js_cert['certificate'] = []
                  js_certificate = { "certificate": {}, "self_signed": {}, "common_name": {}, "subject_alt_names": {} }
                  # Hardcode some values as no time to update csv
                  # Add to dict
                  js_certificate['certificate'] = d['ssl_key_and_certificate_refs']
                  js_certificate['self_signed'] = 'false'
                  js_certificate['common_name'] = d['fqdn']
                  if d['gslb_cname'].lower() == 'none' :
                      js_certificate['subject_alt_names'] = str(d['fqdn'])
                  else:                      
                      js_certificate['subject_alt_names'] = '['+str(d['fqdn'])+ ', ' + str(d['gslb_cname']) + ']'
                  js_cert['certificate'].append(js_certificate)
                  js_cert['type'] = 'SSL_CERTIFICATE_TYPE_VIRTUALSERVICE'
                  if current_row and last_row != d['vs_name']:
                     dict_per_app.append(js_current)        # Write out object to arai
                  js_current = js_cert # Set newly created object to spare variable
                  current_row = d['vs_name']
            elif current_row == d['vs_name']:
                last_row = d['vs_name'] 
        else:
            continue
    dict_per_app.append(js_current) # Dump last row
    return dict_per_app

# Create Playbook From List of Apps
def create_playbook_dict_from_app_directories(app_dict):
    #p = pathlib.Path.cwd()
    #parent = p.resolve().parent
    #app_base_path = parent.joinpath("applications")
    #app_base_path = p.joinpath("applications")

    # Get list of Applications that have been generated from the CSV file
    # Although this is cool, we should be using the list as produced via the csv, because the ./applications dir could contain config.yml files for multiple environments
    #app_dirs = sub_dir_path(app_base_path)
    #for d in app_dirs:
    #    app_list.append(d.stem)

    app_list = []
    for d in app_dict['app_dir']:
        app_list.append(d['vs_name'])

    #current_tenant = ""
    #next_tenant = ""
    #for t in app_dict['app_dir']:
    #    if current_tenant != t['tenant']:
    #        tenant = t['tenant']
    #        if tenant and next_tenant != t['tenant']:
    #            sys.exit()
    #        current_tenant = tenant
    #    elif current_tenant == t['tenant']:
    #        next_tenant = t['tenant']

    tenant = ""
    for t in app_dict['app_dir']:    # This needs a check
        tenant = t['tenant']

    deployment_target = ""
    for dt in app_dict['app_dir']:
        deployment_target = dt['deployment_target']

    if len(app_list) == 0:
        logger.warning("Application list is empty")
        sys.exit()

    # Create the Playbook YAML
    root = [] # Not sure about this is required
    js = {}
    js['hosts'] = "localhost"
    js['connection'] = "local"
    js['gather_facts'] = False

    js_vars = {}
    js_vars['deployment_target'] = deployment_target # "prod_a" # Need to provide via variable
    js_vars['tenant'] = tenant # "BSS_ProdA"         # Need to provide via variable
    js['vars'] = js_vars

    js_roles_list = []
    js_roles = {}
    js_roles['role'] = "avinetworks.avisdk"
    js_roles['tags'] = ['always', 'dbug']
    js_roles_list.append(js_roles)
    js['roles'] = js_roles_list

    # Application List Task
    js_tasks_list = []
    js_tasks = {}
    js_tasks['name'] = "Set Application var Config Directory"
    js_tasks['tags'] = ['always', 'dbug']
    js_app_list = {}
    js_app_list['app_list'] = "{{ app_list }} + [\u0027{{ site_dir }}/applications/{{ tenant }}/{{ item }}/config.yml\u0027]"
    js_tasks['set_fact'] = js_app_list

    js_tasks_vars = {}
    js_tasks_app_list = []
    js_tasks_vars['app_list'] = js_tasks_app_list
    js_tasks_vars['apps'] = app_list # As returned from Function above ['app1', 'app2', 'app3', 'app4']
    js_tasks['vars'] = js_tasks_vars

    js_tasks_with_items_list = []
    js_tasks_with_items_list.append("{{ apps }}")
    js_tasks['with_items'] = js_tasks_with_items_list 

    js_tasks_list.append(js_tasks)

    # Debug Task
    js_tasks = {}
    js_tasks['name'] = "Debug Full avi_config_file passed filename"
    js_tasks['tags'] = ['never', 'dbug']

    js_task_debug = {}
    js_task_debug['msg'] = "{{ item }}"
    js_task_debug['verbosity'] = 1
    js_tasks['debug'] = js_task_debug

    js_tasks_with_items_list = []
    js_tasks_with_items_list.append("{{ app_list }}")
    js_tasks['with_items'] = js_tasks_with_items_list 

    js_tasks_list.append(js_tasks)

    # Final Task
    js_tasks = {}
    js_tasks['name'] = "Avi Application | Setup App"

    js_tasks_incude_role = {}
    js_tasks_incude_role['name'] = "avinetworks.aviconfig"
    js_tasks['include_role'] = js_tasks_incude_role

    js_tasks_vars = {}
    js_tasks_vars['avi_config_file'] = "{{ app_list }}"
    js_tasks_vars['avi_creds_file'] = "{{ site_dir }}/creds/{{deployment_target | default('dev')}}/creds.yml"
    js_tasks['vars'] = js_tasks_vars

    js_tasks_with_items_list = []
    js_tasks_with_items_list.append("{{ app_list }}")
    js_tasks['with_items'] = js_tasks_with_items_list 

    js_tasks_loop_control = {}
    js_tasks_loop_control['loop_var'] = "avi_config_file"
    js_tasks['loop_control'] = js_tasks_loop_control

    js_tasks_list.append(js_tasks)

    # Finally Add List to Stanza
    js['tasks'] = js_tasks_list 
    root.append(js)    
    return root

def create_prerequisites(app_dict):
    app_list = []
    tenant = ""
    deployment_target = ""
    for d in app_dict['app_dir']:
        app_list.append(d['vs_name'])
        tenant = d['tenant']
        deployment_target = d['deployment_target']

    if len(app_list) == 0:
        logger.warning("Application list is empty")
        sys.exit()

    # Create the Playbook YAML

    # Application List Task   
    yml_preqs = [{'connection': 'local',
  'gather_facts': False,
  'hosts': 'localhost',
  'roles': [{'role': 'avinetworks.avisdk'}],
  'tasks': [{'name': 'Create prerequisites',
             'set_fact': {'avi_config': {'applicationpersistenceprofile': [{'app_cookie_persistence_profile': {'prst_hdr_name': 'JSESSIONID',
                                                                                                               'timeout': 20},
                                                                            'description': 'JSESSIONID cookie persistence',
                                                                            'name': 'JSESSIONID-Cookie',
                                                                            'persistence_type': 'PERSISTENCE_TYPE_APP_COOKIE',
                                                                            'state': 'present',
                                                                            'tenant': '{{ tenant }}'}],
                                         'vsdatascriptset': [{'datascript': [{'evt': 'VS_DATASCRIPT_EVT_HTTP_REQ',
                                                                              'script': 'host = avi.http.get_header("Host")\nport = avi.vs.port()\nprotocol = avi.http.protocol()\nif protocol and port then\n    avi.http.add_header( \'Forwarded\', host..\':\'..port )\n    avi.http.add_header( \'proto\', protocol )\n    avi.http.add_header( \'WL-Proxy-SSL\', \'true\' )\nend'}],
                                                              'name': '{{ dsname }}',
                                                              'state': 'present',
                                                              'tenant': '{{ tenant }}'}]}}},
            {'debug': {'msg': '{{ avi_config }}', 'verbosity': 1},
             'name': 'Debug Passed Vars',
             'tags': 'dbug'},
            {'import_role': {'name': 'avinetworks.aviconfig'},
             'name': 'Avi Application | Setup foo',
             'tags': ['always'],
             'vars': {'avi_config': '{{ avi_config }}',
                      'avi_creds_file': '{{ site_dir }}/creds/{{deployment_target}}/creds.yml'}}],
  'vars': {'deployment_target': deployment_target,
           'dsname': 'add-headers',
           'tenant': tenant}}]
    return yml_preqs

# Workhorse Function to Return dictionary path for a particular passed value as a list
def gen_dict_extract_path(var, to_find, curr_pos = None):
        
    if curr_pos is None:
        curr_pos = [] # list to hold key path
    if hasattr(var,'items'):
        for k, v in var.items():
            curr_pos.append(k)
            if to_find in [v]:          
                yield v,curr_pos[:]
            elif to_find in str([v]) and isinstance(v, str):
                yield v,curr_pos[:]            
            if isinstance(v, dict):
                for result in gen_dict_extract_path(v, to_find, curr_pos):
                    yield result
            elif isinstance(v, list):
                curr_pos.append(0)
                for ind,d in enumerate(v):
                    curr_pos.pop()
                    curr_pos.append(ind)
                    for result in gen_dict_extract_path(d, to_find, curr_pos):
                        yield result
                curr_pos.pop()
            curr_pos.pop()
                    
    elif isinstance(var, list):
        curr_pos.append(0)
        for ind,d in enumerate(var):
            curr_pos.pop()
            curr_pos.append(ind)
            for result in gen_dict_extract_path(d, to_find, curr_pos):
                yield result
        curr_pos.pop()
    elif isinstance(var, str):      # If var is a string value as part of a list[], search the string for the value None
        if to_find in var:
            yield var,curr_pos[:-1] # Return the key -1 and not the list index, because otherwise you'll end up with an empty list

# Return Key Path given key
def get_dict_path_from_key(var, key, curr_pos = None):
  """
  key: key to search for
  var: nested dict to search in 
  """
  if curr_pos is None:
    curr_pos=[]
  if hasattr(var,'items'):
    for k, v in var.items():
      curr_pos.append(k)
      if k == key:
        yield v,curr_pos[:]
      if isinstance(v, dict):
        for result in get_dict_path_from_key(v, key, curr_pos):
          yield result
      elif isinstance(v, list):
        curr_pos.append(0)
        for ind,d in enumerate(v):
          curr_pos.pop()
          curr_pos.append(ind)
          for result in get_dict_path_from_key(d, key, curr_pos):
            yield result
        curr_pos.pop()
      curr_pos.pop()
  elif isinstance(var, list):
    curr_pos.append(0)
    for ind,d in enumerate(var):
      curr_pos.pop()
      curr_pos.append(ind)
      for result in get_dict_path_from_key(d, key, curr_pos):
        yield result
    curr_pos.pop()

# Function to return a list of keys, where 'None' as a keyword is detected
def return_key_from_value_list(dic, value):
    key_list = []
    for r in list(gen_dict_extract_path(dic, value)):   # Get a list of paths for all occurances of 'None' using function above      
        logger.debug("Found Key: %s", r[1])
        key_list.append(r[1])                           # This creates a tuple, so only adding the second element (the key represented as a list) to the key_list
    return key_list
        

# Function to update a set of keys for a given Value
def return_key_path_from_given_key_list(dic, value):
    key_list = []
    for r in list(get_dict_path_from_key(dic, value)):   # Get a list of paths for all occurances of 'None' using function above
        key_list.append(r[1])  # This creates a tuple, so only adding the second element (the key represented as a list) to the key_list
    return key_list    

# Used for Testing - get_nested_value_from_key_submitted_as_string
def get_nested_value(dict, list):
    length = len(list)
    try:
        for depth, key in enumerate(list):
            if depth == length - 1:
                output = dict[key]
                return output
            dict = dict[key]
    except (KeyError, TypeError):
        return None
    return None

# Update Value for Given Key
def update_nested_key_value(dict, list, value):
    # update_nested_key_value(js, key_list[1], 'absent')
    length = len(list)
    try:
        for depth, key in enumerate(list):
            if depth == length - 1:
                dict[key] = value
            dict = dict[key]
    except (KeyError, TypeError):
        return None
    return None

# Remove the Discovered nested Keys from the Dict(js)
def delete_nested_key(dict, list, key_to_delete):
    length = len(list)
    try:
        for depth, key in enumerate(list):
            if depth == length - 1:
                dict[key].pop(key_to_delete) # Delete the key from the Dictionary
                logger.debug("Deleting Key: %s, %s" ,list, key_to_delete)
            dict = dict[key]
    except (KeyError, TypeError):
        return None
    return None

# Retrieve key list where value == 'None' and strip the occurances from the dict
def delete_keys_with_value(dic, value):
    key_list = return_key_from_value_list(dic, value)
    for k in key_list:
        key_to_delete = k.pop() # Remember to re-run key_list for loop as this removes values from the array
        delete_nested_key(dic, k, key_to_delete)

# Quickly Modify Values on the fly without regenerating yaml
def update_all_keys_with_value(dic, key, value):
    # update_all_keys_with_value(js, 'state', 'absent')
    key_list = return_key_path_from_given_key_list(dic, key)
    for k in key_list:
        update_nested_key_value(dic, k, value)

# Pretty Print some YAML
def pretty(data):
    print(yaml.dump(json.loads(str(json.dumps(data))), sort_keys=False, default_flow_style=False, encoding = None, explicit_start = True, default_style = '' )) #encoding=('utf-8')

def sub_dir_path (d):
    return [f for f in d.iterdir() if f.is_dir()]

# Create Directories for Each Processed Application
def create_scaffold_dirs(csvFile):
    #csvSeek = codecs.open(csvFile, "r", 'utf-8')
    #csvOpen = csv.reader(csvSeek)
    #csvSeek.seek(0)
    #keys = next(csvOpen)
    LBSheet = csvFile["LB"]
    max_col = LBSheet.max_column
    max_row = LBSheet.max_row
    keys = []
    for i in range(1, max_col + 1):
      cell_obj = LBSheet.cell(row = 3, column = i)
      key= cell_obj.value
      keys.append(key)
    fullrow_value= []
    for x in range (4, max_row + 1):
      onerow_value = []
      if LBSheet.cell(row = x, column = 1).value:
        for i in range(1, max_col + 1):
          cell_obj = LBSheet.cell(row = x, column = i)
          cell_value= cell_obj.value
          onerow_value.append(cell_value)
        fullrow_value.append(onerow_value)
    js_dirs = {}
    js_dirs['app_dir'] = []
    currentRow = None
    for row in fullrow_value:
        d = dict(zip(keys, row))
        create=d.get('create', 'false')
        if str(create).lower() == 'true' or str(create).lower() == 'yes':
            if currentRow != d['vs_name'] and str(d['create']).lower() != 'false': # vs_name will always be unique based on tenant-dc-app-name
                js_app_dir = {}
                js_app_dir['vs_name'] = d['vs_name']
                js_app_dir['tenant'] = d['tenant']
                currentRow = d['vs_name']
                js_dirs['app_dir'].append(js_app_dir)
            else:
                continue
        else:
            continue

    app_dirs_arai = []
    p = pathlib.Path(__file__).resolve().parent.parent
    parent = p.resolve()
    app_base_path = parent.joinpath("applications")
    for dir in js_dirs['app_dir']:
        app_dir = dir['vs_name']
        app_specific_path = app_base_path.joinpath(dir['tenant']) #testtttttttt
        if (app_specific_path.exists() == False):
            app_specific_path.mkdir(parents=True, exist_ok=True) # mkdir
        else:
            logger.info("Directory: %s exists", dir)
        app_dirs_arai.append(app_specific_path)
    return app_dirs_arai

def application_directories(csvFile):
    LBSheet = csvFile["LB"]
    max_col = LBSheet.max_column
    max_row = LBSheet.max_row
    keys = []
    for i in range(1, max_col + 1):
      cell_obj = LBSheet.cell(row = 3, column = i)
      key= cell_obj.value
      keys.append(key)
    fullrow_value= []
    for x in range (4, max_row + 1):
      onerow_value = []
      if LBSheet.cell(row = x, column = 1).value:
        for i in range(1, max_col + 1):
          cell_obj = LBSheet.cell(row = x, column = i)
          cell_value= cell_obj.value
          onerow_value.append(cell_value)
        fullrow_value.append(onerow_value)

    #csvSeek = codecs.open(csvFile, "r", 'utf-8')
    #csvOpen = csv.reader(csvSeek)
    #keys = next(csvOpen)
    js_dirs = {}
    js_dirs['app_dir'] = []
    currentRow = None
    for row in fullrow_value:
        d = dict(zip(keys, row))
        if currentRow != d['vs_name']: # vs_name will always be unique based on tenant-dc-app-name
            js_app_dir = {}
            js_app_dir['vs_name'] = d['vs_name']
            logger.info(d['vs_name'])
            js_app_dir['tenant'] = d['tenant']
            deployment_target = d['dc']
            if deployment_target == "a":
                js_app_dir['deployment_target'] = "prod_" + d['dc']
            elif deployment_target == "b":
                js_app_dir['deployment_target'] = "prod_" + d['dc']
            elif deployment_target == "dev_a":
                js_app_dir['deployment_target'] = d['dc']
            elif deployment_target == "dev_b":
                js_app_dir['deployment_target'] = d['dc']
            elif deployment_target == "preprod_a":
                js_app_dir['deployment_target'] = d['dc']
            elif deployment_target == "preprod_b":
                js_app_dir['deployment_target'] = d['dc']

            currentRow = d['vs_name']
            #app_directories.append(currentRow)
            if str(d['create']).lower() == 'true':
                js_dirs['app_dir'].append(js_app_dir)
    return js_dirs

# Write out each processed dictionary as individual yaml files into their relevant directories
#def append_multiple_lines(output_path, output):
 ##    with open(output_path, "a+") as yaml_file:        #testtttttttt
  #    appendEOL = False
  #    yaml_file.seek(0)
   #   data = yaml_file.read(1000)
   #   if len(data) > 0 :
   #    appendEOL = True
   #   for lines in yaml_file.readlines() : 
   #    for line in output:
   #     if appendEOL == True:
   #       yaml_file.write("\n")
   #     else:
   #       appendEOL = True
   #     yaml_file.write(line)                          
   #    #yaml_file.close()
def write_to_yaml(data, app_dir, output_filename):
    yaml.preserve_quotes = True
    output = yaml.dump(json.loads(str(json.dumps(data))), sort_keys=False, default_flow_style=False, encoding = None, explicit_start = True, default_style = '')
    #output_path = app_dir.joinpath("config.yml")
    output_path = app_dir.joinpath(output_filename)
   # append_multiple_lines(output_path, output)                           
    yaml_file = open(output_path, "w")
    yaml_file.write(output)
    yaml_file.close()

'''
def output_playbook(csvFile, args):
    p = pathlib.Path(__file__).resolve().parent.parent
    output_path = p
    # .resolve().parent
    # output_path = 'applications'+app_dict['app_dir'][0]['tenant']+'/'
    app_dict = application_directories(csvFile)
    if args.playbook_file_prefix:
        output_filename = app_dict['app_dir'][0]['deployment_target'].lower() + "_" + app_dict['app_dir'][0]['tenant'].lower() + ".yml"
        prereq_filename = app_dict['app_dir'][0]['deployment_target'].lower() + "_" + app_dict['app_dir'][0]['tenant'].lower() + "_prereqs.yml"
        sgs_play_filename = app_dict['app_dir'][0]['deployment_target'].lower() + "_" + app_dict['app_dir'][0]['tenant'].lower() + "_sgs.yml"
    else:
        output_filename = 'applications/'+app_dict['app_dir'][0]['tenant']+'/'+"nogit_" + app_dict['app_dir'][0]['deployment_target'].lower() + "_" + app_dict['app_dir'][0]['tenant'].lower() + ".yml"
        prereq_filename = 'applications/'+app_dict['app_dir'][0]['tenant']+'/'+"nogit_" + app_dict['app_dir'][0]['deployment_target'].lower() + "_" + app_dict['app_dir'][0]['tenant'].lower() + "_prereqs.yml"
        sgs_play_filename = 'applications/'+app_dict['app_dir'][0]['tenant']+'/'+"nogit_" + app_dict['app_dir'][0]['deployment_target'].lower() + "_" + app_dict['app_dir'][0]['tenant'].lower() + "_sgs.yml"

    playbook_dict = create_playbook_dict_from_app_directories(app_dict) # Update to pass in tenant
    prerequisite_dict = create_prerequisites(app_dict)
    d_sgs = funks.create_security_groups_playbook(app_dict)

    if args.print_to_console:
        pretty(playbook_dict)
    elif args.output == "sgs":
        sgs_play_filename = 'applications/'+app_dict['app_dir'][0]['tenant']+'/'+"nogit_" + app_dict['app_dir'][0]['deployment_target'].lower() + "_" + app_dict['app_dir'][0]['tenant'].lower() + "_sgs.yml"
        write_to_yaml(d_sgs, output_path, sgs_play_filename)
    else:
        write_to_yaml(playbook_dict, output_path, output_filename)
        write_to_yaml(prerequisite_dict, output_path, prereq_filename)
        write_to_yaml(d_sgs, output_path, sgs_play_filename)
        print("For prerequisites:")
        print("export PLAY=" + str(output_path) +'/'+ prereq_filename)
        print("For the main playbook:")
        print("export PLAY=" + str(output_path) +'/'+ output_filename)
        print("To run the playbook:")
        print("ansible-playbook -v $PLAY --extra-vars \"site_dir='"+str(output_path)+"'\" --vault-id creds/.vault_password")
        print("For Security Groups:")
        print("export PLAY=" + str(output_path) +'/'+ sgs_play_filename)
        print("ansible-playbook $PLAY --extra-vars \"site_dir='"+str(output_path)+"'\"")
'''
def main():

    parser = argparse.ArgumentParser(description = 'CSV to YAML Converter for SLB Input File', formatter_class = argparse.RawTextHelpFormatter)
    parser.add_argument('csv_input', type = str, help = "Specify input CSV file as file.csv")
    parser.add_argument('--output', dest = 'output', required = False, default = "all", choices=['all', 'sgs', 'csr_json', 'playbook'], 
        help = textwrap.dedent('''\
        Specify desired output config: "default: all"
        \tall - All config is generated
        \tsgs - Security Groups Only
        \tcsr_json - Only JSON for Creating CSR's
        \tplaybook - Only required playbooks; prereqs playbook, main playbook, security group playbook
        '''))
    parser.add_argument('--print', dest = 'print_to_console', action = 'store_true', help = "Only Print YAML to Console")
    parser.add_argument('--no-prefix', dest = 'playbook_file_prefix', action = 'store_true', help = "Provide Prefix for Output Playbook File")
    parser.add_argument('--no-key-delete', dest = 'no_delete', action = 'store_true', help = "Don't delete keys of None Type")
    parser.add_argument('-v', '--verbose', dest = 'log_level', action='count', default = 0, help = "Set Output level. Use -vvv to increase logging level")
    args = parser.parse_args()
    

    # Set Logging level based on verbosity specified
    if args.log_level == 1:
        logger.setLevel(logging.WARNING)
    elif args.log_level == 2:
        logger.setLevel(logging.INFO)
    elif args.log_level == 3:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.WARNING) # default is 0, so only print log messages @ WARNING and below (CRITICAL) level, and nothing above it, e.g. INFO

    p = pathlib.Path.cwd()
    csvstatus = p.joinpath(args.csv_input)
    csvFile = load_workbook(args.csv_input, data_only=True)
    
    
    if (csvstatus.exists() == False):
        logger.critical("File: %s doesn't exist", csvstatus)
        sys.exit()
    validate_csv_file(csvFile)    
    # Generate dictionaries from CSV
    pools_arai = create_pools_lod_from_csv(csvFile)
    virtual_services_arai = create_vs_lod_from_csv(csvFile)
    ssl_certificate_arai = create_cert_load_from_csv(csvFile)
    ipaddr_groups_arai = funks.create_ipaddr_groups_from_csv(csvFile)   #  Pull back all ipaddrgroup data
    httppolicyset_arai = funks.create_httppolicyset_from_csv(csvFile)
    d_security_group_vars = funks.create_sg_vars_from_csv(csvFile)

    # Create Directories if --print is not specified
    if args.print_to_console != True:
        app_dirs = create_scaffold_dirs(csvFile) # Create Folders, 1 per app

    if args.output == "playbook":
        output_playbook(csvFile, args)
    elif args.output == "csr_json":
        p = pathlib.Path.cwd()
        output_path = p #.resolve().parent
        csr_dict = funks.create_csr_json_from_csv(csvFile)
        output_filename = "nogit_csr_input.json"
        if args.print_to_console:
            funks.pretty_json(csr_dict)
        else:
            funks.write_json_to_file(csr_dict, output_path, output_filename)
    elif args.output == "sgs":
        output_playbook(csvFile, args)
        for i in range(len(virtual_services_arai)):
            for j in range(len(d_security_group_vars)):
                for k in range(len(d_security_group_vars[j]['tenant_networks']['security_groups'])):
                    if virtual_services_arai[i]['name'] in d_security_group_vars[j]['tenant_networks']['security_groups'][k]['name']:
                        write_to_yaml(d_security_group_vars[j], app_dirs[i], output_filename = "security_groups.yml")                
    
    else:
        #output_playbook(csvFile, args)
        all_data= {}
        all_data['avi_config'] = []
        
        all_data['avi_security_groups'] = []
        js = {}
        js['pool'] = []
        for i in range(len(virtual_services_arai)):
            # Create final dictionary stanza
            if args.output == "all":
                js['pool'].append(pools_arai[i])
                if bool(ipaddr_groups_arai) == True and len(ipaddr_groups_arai) > 1:
                    for j in range(len(ipaddr_groups_arai)):
                        if virtual_services_arai[i]['name'] in ipaddr_groups_arai[j]['vs_name']:
                                js['ipaddrgroup'] = []
                                js['ipaddrgroup'].append(ipaddr_groups_arai[j]['js_list'])
                else:
                    logger.debug("ipaddr_groups_arai returned empty, is ipaddrgroup_name == None ?")
                if bool(httppolicyset_arai) == True and len(httppolicyset_arai) > 1:
                    for j in range(len(httppolicyset_arai)):
                        if virtual_services_arai[i]['name'] in httppolicyset_arai[j]['vs_name']:
                            js['httppolicyset'] = []
                            js['httppolicyset'].append(httppolicyset_arai[j]['js_list'])
                else:
                    logger.info("httppolicyset_arai returned empty, is httppolicyset_name == None ?")
        
        if args.no_delete:
                logger.info(" =| For Testing only |=: No 'None' Type Key Delete - Config WILL be invalid")
        else:
                delete_keys_with_value(js, 'None')
        all_data['avi_config'].append(js)
        js1 = {}
        js1['virtualservice'] = []
        for i in range(len(virtual_services_arai)):
            if args.output == "all":
                js1['virtualservice'].append(virtual_services_arai[i])
        if args.no_delete:
                logger.info(" =| For Testing only |=: No 'None' Type Key Delete - Config WILL be invalid")
        else:
                delete_keys_with_value(js1, 'None')
        all_data['avi_config'].append(js1)
        js2 = {}
        for i in range(len(virtual_services_arai)):
            if args.output == "all":       
                for j in range(len(d_security_group_vars)):

                    for k in range(len(d_security_group_vars[j])):
                        if virtual_services_arai[i]['name'] in d_security_group_vars[j]['security_groups'][k]['name']:
                            
                            all_data['avi_security_groups'].extend(d_security_group_vars[j]['security_groups'])

            # Delete 'None' type key-values from the Final Dictionary
            if args.no_delete:
                logger.info(" =| For Testing only |=: No 'None' Type Key Delete - Config WILL be invalid")
            else:
                delete_keys_with_value(js2, 'None')
        js3 = {}
        js3['sslkeyandcertificate'] = []
        for i in range(len(virtual_services_arai)):
            if args.output == "all":
                for j in range(len(ssl_certificate_arai)):
                    if 'ssl_key_and_certificate_refs' in virtual_services_arai[i].keys():
                      if virtual_services_arai[i]['ssl_key_and_certificate_refs'][31:] in ssl_certificate_arai[j]['name']:
                        js3['sslkeyandcertificate'].append(ssl_certificate_arai[j])
        if args.no_delete:
                logger.info(" =| For Testing only |=: No 'None' Type Key Delete - Config WILL be invalid")
        else:
                delete_keys_with_value(js3, 'None')
        
        all_data['avi_config'].append(js3)
        

            # If specified, print to console, else write to files
        #if args.print_to_console:
        #    pretty(js)
        #    pretty(js1)
        #else:
        #    all_data['avi_config'].append(js)
        #    all_data['avi_config'].append(js1)
     
    env_dict = application_directories(csvFile)
    outfile = str(env_dict['app_dir'][0]['deployment_target']).lower()    
    write_to_yaml(all_data, app_dirs[i], output_filename = 'tenant-avi-config_'+outfile+'.yml')
    

if __name__== "__main__":
    main()

# generate data for ipaddr_groups
# when selecting 'all' search through ipaddr_groups based on ['vs_name']
# set the key to "None" and delete that stanza
